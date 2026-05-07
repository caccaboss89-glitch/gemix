#!/usr/bin/env python3
"""Convert between csv / xlsx / pdf.

Subcommands:
  csv2xlsx  : import a CSV/TSV file into a single-sheet xlsx
  xlsx2csv  : export one sheet (or every sheet) of an xlsx to CSV
  xlsx2pdf  : render the workbook to a PDF via headless LibreOffice

CSV delimiter auto-detection covers ",", ";", "\\t", "|".
"""
import argparse
import csv
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook


_DELIM_CANDIDATES = (",", ";", "\t", "|")


def _detect_delimiter(sample: str, fallback: str = ",") -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(_DELIM_CANDIDATES))
        return dialect.delimiter
    except csv.Error:
        # Fallback heuristic: most occurrences in the first lines.
        counts = {d: sample.count(d) for d in _DELIM_CANDIDATES}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else fallback


def csv2xlsx(input_path: Path, output: Path, delimiter: Optional[str],
             sheet_name: str, has_header: bool) -> None:
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    if input_path.suffix.lower() not in (".csv", ".tsv", ".txt"):
        raise ValueError(f"Unsupported CSV extension: {input_path.suffix}")

    text = input_path.read_text(encoding="utf-8-sig", errors="replace")
    if delimiter is None:
        sample = "\n".join(text.splitlines()[:20])
        delimiter = _detect_delimiter(sample, fallback="\t" if input_path.suffix.lower() == ".tsv" else ",")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Data"

    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    for r_idx, row in enumerate(reader, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            # Preserve formulas if the user explicitly wrote them in the CSV.
            if isinstance(val, str) and val.startswith("="):
                cell.value = val
            else:
                # Try numeric coercion (most CSVs round-trip via openpyxl as text otherwise).
                try:
                    if isinstance(val, str) and val.strip() == "":
                        cell.value = None
                    elif isinstance(val, str) and ("." in val or "e" in val.lower()):
                        cell.value = float(val)
                    elif isinstance(val, str):
                        cell.value = int(val)
                    else:
                        cell.value = val
                except ValueError:
                    cell.value = val
    if has_header and ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        ws.freeze_panes = "A2"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Converted -> {output} (delimiter='{repr(delimiter)[1:-1]}', rows={ws.max_row}, cols={ws.max_column})")


def _write_sheet_to_csv(ws, out_path: Path, delimiter: str) -> int:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        rows = 0
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            rows += 1
    return rows


def xlsx2csv(input_path: Path, sheet: Optional[str], all_sheets: bool,
             output: Optional[Path], output_dir: Optional[Path],
             delimiter: str, data_only: bool) -> None:
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported extension: {input_path.suffix}")

    wb = load_workbook(input_path, data_only=data_only)

    if all_sheets:
        if not output_dir:
            raise ValueError("--all requires --output-dir")
        for name in wb.sheetnames:
            safe = "".join(c for c in name if c.isalnum() or c in "-_") or "sheet"
            target = output_dir / f"{safe}.csv"
            n = _write_sheet_to_csv(wb[name], target, delimiter)
            print(f"Sheet '{name}' -> {target}  ({n} rows)")
        return

    name = sheet or wb.sheetnames[0]
    if name not in wb.sheetnames:
        raise ValueError(f"Sheet '{name}' not found. Available: {wb.sheetnames}")
    if not output:
        raise ValueError("--output is required unless --all is used")
    n = _write_sheet_to_csv(wb[name], output, delimiter)
    print(f"Sheet '{name}' -> {output}  ({n} rows)")


def xlsx2pdf(input_path: Path, output: Path, timeout: int) -> None:
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("LibreOffice (soffice) not found in PATH.")

    with tempfile.TemporaryDirectory(prefix="lo_pdf_") as tmpdir:
        tmp = Path(tmpdir)
        profile_dir = tmp / f"profile_{uuid.uuid4().hex[:8]}"
        out_dir = tmp / "out"
        out_dir.mkdir()
        profile_uri = profile_dir.absolute().as_uri()

        cmd = [
            soffice,
            f"-env:UserInstallation={profile_uri}",
            "--headless",
            "--calc",
            "--norestore",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to", "pdf",
            "--outdir", str(out_dir),
            str(input_path),
        ]
        try:
            res = subprocess.run(
                cmd, capture_output=True, text=True, timeout=timeout,
                env={**os.environ, "HOME": str(profile_dir)},
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(f"xlsx2pdf timed out after {timeout}s") from exc
        if res.returncode != 0:
            raise RuntimeError(
                f"soffice exited {res.returncode}.\nstdout: {res.stdout}\nstderr: {res.stderr}"
            )

        produced = out_dir / (input_path.stem + ".pdf")
        if not produced.exists():
            pdfs = list(out_dir.glob("*.pdf"))
            if not pdfs:
                raise RuntimeError(f"soffice produced no PDF. stderr: {res.stderr}")
            produced = pdfs[0]

        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(produced, output)
        print(f"Converted -> {output}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert between csv / xlsx / pdf.")
    sub = parser.add_subparsers(dest="action", required=True)

    c = sub.add_parser("csv2xlsx", help="CSV/TSV → XLSX")
    c.add_argument("--input", required=True)
    c.add_argument("--output", required=True)
    c.add_argument("--delimiter", default=None,
                   help="Delimiter (default: auto-detect , ; tab |)")
    c.add_argument("--sheet-name", default="Data")
    c.add_argument("--no-header", action="store_true",
                   help="Treat first row as data (no bold/freeze)")

    x = sub.add_parser("xlsx2csv", help="XLSX → CSV")
    x.add_argument("--input", required=True)
    x.add_argument("--sheet", help="Sheet name (default: first sheet)")
    x.add_argument("--all", action="store_true", help="Export every sheet (requires --output-dir)")
    x.add_argument("--output", help="Output CSV path (single sheet mode)")
    x.add_argument("--output-dir", help="Output directory (--all mode)")
    x.add_argument("--delimiter", default=",", help="Output delimiter (default: ',')")
    x.add_argument("--no-data-only", action="store_true",
                   help="Export raw formula strings instead of cached values")

    p = sub.add_parser("xlsx2pdf", help="XLSX → PDF (LibreOffice headless)")
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--timeout", type=int, default=90)

    args = parser.parse_args()
    try:
        if args.action == "csv2xlsx":
            csv2xlsx(
                Path(args.input),
                Path(args.output),
                args.delimiter,
                args.sheet_name,
                has_header=not args.no_header,
            )
        elif args.action == "xlsx2csv":
            xlsx2csv(
                Path(args.input),
                sheet=args.sheet,
                all_sheets=args.all,
                output=Path(args.output) if args.output else None,
                output_dir=Path(args.output_dir) if args.output_dir else None,
                delimiter=args.delimiter.replace("\\t", "\t"),
                data_only=not args.no_data_only,
            )
        elif args.action == "xlsx2pdf":
            xlsx2pdf(Path(args.input), Path(args.output), timeout=args.timeout)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
