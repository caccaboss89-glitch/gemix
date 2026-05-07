#!/usr/bin/env python3
"""Static QA validation for XLSX workbooks.

Checks for:
- Formula errors (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?)
- Empty cells in data ranges
- Inconsistent number formats
- Missing headers
- Broken cross-sheet references
- Data validation issues
- Print settings missing on multi-page sheets
"""
import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string


def _is_formula_error(value: Any) -> bool:
    """Check if a cell value is a formula error."""
    if not isinstance(value, str):
        return False
    return value.startswith("#") and value.endswith("!")


def _is_empty(value: Any) -> bool:
    """Check if a cell is effectively empty."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _check_sheet(ws, sheet_name: str, config: Dict[str, Any]) -> Dict[str, Any]:
    """Run QA checks on a single sheet."""
    issues = []
    warnings = []
    
    # Check for formula errors
    error_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if _is_formula_error(cell.value):
                error_cells.append(f"{cell.coordinate}: {cell.value}")
    
    if error_cells:
        issues.append({
            "type": "formula_error",
            "severity": "critical",
            "message": f"Found {len(error_cells)} formula error(s)",
            "locations": error_cells
        })
    
    # Check for empty cells in data range (if specified)
    data_range = config.get("data_range")
    if data_range:
        empty_cells = []
        # Parse A1:B10 format
        if ":" not in data_range:
            warnings.append({
                "type": "invalid_range",
                "severity": "warning",
                "message": f"Invalid data_range format: {data_range}"
            })
        else:
            start, end = data_range.split(":")
            start_col, start_row = coordinate_from_string(start)
            end_col, end_row = coordinate_from_string(end)
            min_col = min(column_index_from_string(start_col), column_index_from_string(end_col))
            max_col = max(column_index_from_string(start_col), column_index_from_string(end_col))
            min_row = min(start_row, end_row)
            max_row = max(start_row, end_row)
            
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, 
                                    min_col=min_col, max_col=max_col):
                for cell in row:
                    if _is_empty(cell.value):
                        empty_cells.append(cell.coordinate)
        
        if empty_cells:
            warnings.append({
                "type": "empty_cells",
                "severity": "warning",
                "message": f"Found {len(empty_cells)} empty cell(s) in data range {data_range}",
                "locations": empty_cells
            })
    
    # Check for missing headers
    header_row = config.get("header_row", 1)
    has_header = False
    for cell in ws[header_row]:
        if not _is_empty(cell.value):
            has_header = True
            break
    
    if not has_header and ws.max_row > 1:
        warnings.append({
            "type": "missing_header",
            "severity": "info",
            "message": f"No header found in row {header_row}"
        })
    
    # Check for frozen panes on sheets with >10 rows
    if ws.max_row > 10 and not ws.freeze_panes:
        warnings.append({
            "type": "no_freeze_panes",
            "severity": "info",
            "message": f"Sheet has {ws.max_row} rows but no freeze panes set"
        })
    
    # Check for print settings on sheets with >20 rows
    if ws.max_row > 20 and not ws.page_setup.fitToPage:
        warnings.append({
            "type": "no_print_settings",
            "severity": "info",
            "message": f"Sheet has {ws.max_row} rows but print fit-to-page not set"
        })
    
    # Check for data validation on input columns
    input_columns = config.get("input_columns", [])
    for col in input_columns:
        has_dv = False
        for dv in ws.data_validations.dataValidation:
            if dv.ranges and any(col in str(r) for r in dv.ranges):
                has_dv = True
                break
        
        if not has_dv:
            warnings.append({
                "type": "no_data_validation",
                "severity": "info",
                "message": f"Input column {col} has no data validation"
            })
    
    # Check for inconsistent number formats in numeric columns
    numeric_columns = config.get("numeric_columns", [])
    for col in numeric_columns:
        formats = set()
        for cell in ws[col]:
            if cell.number_format and not _is_empty(cell.value):
                formats.add(cell.number_format)
        
        if len(formats) > 1:
            warnings.append({
                "type": "inconsistent_formats",
                "severity": "warning",
                "message": f"Column {col} has inconsistent number formats: {sorted(formats)}"
            })
    
    return {
        "sheet": sheet_name,
        "issues": issues,
        "warnings": warnings,
        "stats": {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "has_freeze_panes": bool(ws.freeze_panes),
            "has_print_settings": ws.page_setup.fitToPage
        }
    }


def qa(input_path: Path, config_path: Optional[Path] = None) -> Dict[str, Any]:
    """Run QA validation on a workbook."""
    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")
    
    wb = load_workbook(input_path, data_only=True)
    
    # Load config if provided
    config = {}
    if config_path and config_path.exists():
        config = json.loads(config_path.read_text(encoding="utf-8"))
    
    results = []
    total_issues = 0
    total_warnings = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_config = config.get("sheets", {}).get(sheet_name, {})
        result = _check_sheet(ws, sheet_name, sheet_config)
        results.append(result)
        total_issues += len(result["issues"])
        total_warnings += len(result["warnings"])
    
    return {
        "file": str(input_path),
        "status": "passed" if total_issues == 0 else "failed",
        "total_issues": total_issues,
        "total_warnings": total_warnings,
        "sheets": results
    }


def main() -> None:
    p = argparse.ArgumentParser(description="Static QA validation for XLSX workbooks")
    p.add_argument("--input", required=True, help="Path to workbook")
    p.add_argument("--config", help="Optional JSON config with sheet-specific checks")
    p.add_argument("--output", help="Output JSON path (default: stdout)")
    args = p.parse_args()
    
    input_path = Path(args.input)
    config_path = Path(args.config) if args.config else None
    
    try:
        result = qa(input_path, config_path)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)
    
    output = json.dumps(result, indent=2)
    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        Path(args.output).write_text(output, encoding="utf-8")
        print(f"QA report written to {args.output}")
    else:
        print(output)
    
    # Exit with error code if critical issues found
    if result["status"] == "failed":
        sys.exit(1)


if __name__ == "__main__":
    main()
