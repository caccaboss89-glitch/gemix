#!/usr/bin/env python3
"""Build an .xlsx workbook from a JSON spec.

Spec schema (see SKILL.md for the full reference):

{
  "sheets": [
    {
      "name": "Summary",
      "freeze": "A2",
      "auto_width": true,
      "columns": [
        {"col": "A", "width": 22},
        {"col": "B", "number_format": "$#,##0;($#,##0);-"}
      ],
      "rows": [
        {"values": ["Metric", "Value"], "header": true},
        {"values": ["Revenue", {"value": 1200000, "semantic": "input"}]},
        {"values": ["Margin",  {"value": "=B2*0.2", "semantic": "formula"}]}
      ],
      "merges": ["A1:B1"],
      "charts": [
        {"type": "bar", "title": "T", "data_range": "A1:B5",
         "categories_range": "A2:A5", "anchor": "D2"}
      ]
    }
  ],
  "properties": {"title": "Q4 Report", "creator": "GemiX AI"}
}

Cell value forms inside rows[].values[]:
  - Plain scalar (str/int/float/bool/None)
  - Formula string starting with "="
  - Object: {"value": <scalar-or-formula>,
             "semantic": "input|formula|link|external|assumption",
             "number_format": "...", "bold": bool, "fill": "RRGGBB",
             "font_color": "RRGGBB", "alignment": "left|center|right"}
"""
import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, AreaChart
from openpyxl.styles import Alignment, Font, PatternFill, GradientFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule, CellIsRule
from openpyxl.chart.trendline import Trendline


SEMANTIC_FONT_COLOR = {
    "input": "0000FF",
    "formula": "000000",
    "link": "008000",
    "external": "FF0000",
    "assumption": "000000",
}
SEMANTIC_FILL = {
    "assumption": "FFFF00",
}

HEADER_FILL = "D9D9D9"
ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ISO_DATETIME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?$")


def _parse_iso_date(value: str) -> Optional[dt.datetime]:
    try:
        if ISO_DATETIME_RE.match(value):
            return dt.datetime.fromisoformat(value.replace("T", " "))
        if ISO_DATE_RE.match(value):
            return dt.datetime.fromisoformat(value)
    except ValueError:
        return None
    return None


def _is_date_format(fmt: Optional[str]) -> bool:
    if not fmt:
        return False
    f = fmt.lower()
    return any(tok in f for tok in ("yy", "mm", "dd", "hh:mm"))


def _apply_gradient_fill(cell, fill_spec: Dict[str, Any]) -> None:
    """Apply gradient fill to a cell."""
    if not isinstance(fill_spec, dict):
        return
    start_color = fill_spec.get("start_color", "FFFFFF")
    end_color = fill_spec.get("end_color", "000000")
    angle = int(fill_spec.get("angle", 90))
    # openpyxl GradientFill uses 'degree' for linear gradients
    try:
        cell.fill = GradientFill(
            fill_type="linear",
            degree=angle,
            stop=[Color(start_color), Color(end_color)]
        )
    except Exception:
        # Fallback to solid fill if gradient fails
        cell.fill = PatternFill("solid", start_color=start_color, end_color=start_color)


def _apply_border(cell, border_spec: Dict[str, Any]) -> None:
    """Apply border styles to a cell."""
    if not isinstance(border_spec, dict):
        return
    
    def _parse_side(side_spec: Any) -> Optional[Side]:
        if not side_spec:
            return None
        if isinstance(side_spec, str):
            return Side(style=side_spec, color="000000")
        if isinstance(side_spec, dict):
            return Side(
                style=side_spec.get("style", "thin"),
                color=side_spec.get("color", "000000")
            )
        return None
    
    # diagonal_direction is not supported in openpyxl Border, ignore it
    cell.border = Border(
        left=_parse_side(border_spec.get("left")),
        right=_parse_side(border_spec.get("right")),
        top=_parse_side(border_spec.get("top")),
        bottom=_parse_side(border_spec.get("bottom")),
        diagonal=_parse_side(border_spec.get("diagonal"))
    )


def _apply_cell(cell, spec: Any, default_font: Font, column_format: Optional[str]) -> None:
    """Write a value + per-cell styles to an openpyxl Cell."""
    if isinstance(spec, dict) and "value" in spec:
        value = spec.get("value")
        semantic = spec.get("semantic")
        number_format = spec.get("number_format") or column_format
        bold = bool(spec.get("bold", False))
        italic = bool(spec.get("italic", False))
        underline = spec.get("underline")
        fill = spec.get("fill")
        gradient = spec.get("gradient")
        font_color = spec.get("font_color")
        alignment = spec.get("alignment")
        border = spec.get("border")
        comment = spec.get("comment")
        locked = spec.get("locked", True)
    else:
        value = spec
        semantic = None
        number_format = column_format
        bold = False
        italic = False
        underline = None
        fill = None
        gradient = None
        font_color = None
        alignment = None
        border = None
        comment = None
        locked = True

    # ISO date string → datetime, only when the column's number_format is a date format.
    if isinstance(value, str) and _is_date_format(number_format):
        parsed = _parse_iso_date(value)
        if parsed is not None:
            value = parsed

    cell.value = value

    if semantic and font_color is None:
        font_color = SEMANTIC_FONT_COLOR.get(semantic)
    if semantic and fill is None and semantic in SEMANTIC_FILL:
        fill = SEMANTIC_FILL[semantic]

    cell.font = Font(
        name=default_font.name,
        size=default_font.size,
        bold=bold,
        italic=italic,
        underline=underline,
        color=font_color,
    )
    
    if gradient:
        _apply_gradient_fill(cell, gradient)
    elif fill:
        cell.fill = PatternFill("solid", start_color=fill, end_color=fill)
    
    if border:
        _apply_border(cell, border)
    
    if number_format:
        cell.number_format = number_format
    if alignment:
        if isinstance(alignment, str):
            cell.alignment = Alignment(horizontal=alignment)
        elif isinstance(alignment, dict):
            cell.alignment = Alignment(
                horizontal=alignment.get("horizontal"),
                vertical=alignment.get("vertical"),
                wrap_text=alignment.get("wrap_text", False),
                shrink_to_fit=alignment.get("shrink_to_fit", False)
            )
    
    if comment:
        cell.comment = Comment(comment, "GemiX AI")
    
    cell.protection = Protection(locked=bool(locked))


def _write_row(ws, row_idx: int, row_spec: Dict[str, Any], default_font: Font,
               column_formats: Dict[int, str]) -> None:
    values = row_spec.get("values", [])
    is_header = bool(row_spec.get("header", False))
    style_keys = {
        "bold", "italic", "underline", "fill", "gradient", "font_color",
        "alignment", "border", "comment", "locked", "number_format", "semantic"
    }
    for col_idx, raw in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        col_fmt = column_formats.get(col_idx)
        _apply_cell(cell, raw, default_font, col_fmt)
        has_custom_style = isinstance(raw, dict) and any(k in raw for k in style_keys)
        if is_header and not has_custom_style:
            cell.font = Font(name=default_font.name, size=default_font.size, bold=True)
            cell.fill = PatternFill("solid", start_color=HEADER_FILL, end_color=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, max_width: float = 40.0, min_width: float = 10.0) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    text = "12345"  # rough estimate; we don't know the result
                else:
                    text = str(v)
                longest = max(longest, len(text))
        width = max(min_width, min(max_width, longest + 2))
        ws.column_dimensions[letter].width = width


def _resolve_range(rng: str) -> Tuple[int, int, int, int]:
    """Turn "A1:B5" into (min_row, min_col, max_row, max_col)."""
    if ":" not in rng:
        raise ValueError(f"Invalid range '{rng}': expected A1:B5 form.")
    a, b = rng.split(":")
    a_col, a_row = coordinate_from_string(a)
    b_col, b_row = coordinate_from_string(b)
    return (
        min(a_row, b_row),
        min(column_index_from_string(a_col), column_index_from_string(b_col)),
        max(a_row, b_row),
        max(column_index_from_string(a_col), column_index_from_string(b_col)),
    )


def _add_chart(ws, spec: Dict[str, Any]) -> None:
    ctype = spec.get("type", "bar").lower()
    title = spec.get("title")
    data_range = spec.get("data_range")
    cat_range = spec.get("categories_range")
    anchor = spec.get("anchor", "E2")
    if ctype != "combo" and not data_range:
        raise ValueError("chart spec missing 'data_range'")

    secondary_axis = spec.get("secondary_axis", False)

    if ctype in ("bar", "column"):
        chart = BarChart()
        chart.type = "col" if ctype == "column" else "bar"
    elif ctype == "line":
        chart = LineChart()
    elif ctype == "pie":
        chart = PieChart()
    elif ctype == "scatter":
        chart = ScatterChart()
    elif ctype == "area":
        chart = AreaChart()
    elif ctype == "combo":
        primary_type = spec.get("primary_type", "column")
        primary_data = spec.get("primary_data_range")
        secondary_data = spec.get("secondary_data_range")
        if not primary_data:
            raise ValueError("combo chart spec missing 'primary_data_range'")
        
        if primary_type in ("bar", "column"):
            chart = BarChart()
            chart.type = "col" if primary_type == "column" else "bar"
        elif primary_type == "area":
            chart = AreaChart()
        else:
            chart = LineChart()
        
        p_minr, p_minc, p_maxr, p_maxc = _resolve_range(primary_data)
        p_ref = Reference(ws, min_col=p_minc, min_row=p_minr,
                          max_col=p_maxc, max_row=p_maxr)
        chart.add_data(p_ref, titles_from_data=True)
        
        if secondary_data:
            s_minr, s_minc, s_maxr, s_maxc = _resolve_range(secondary_data)
            s_ref = Reference(ws, min_col=s_minc, min_row=s_minr,
                              max_col=s_maxc, max_row=s_maxr)
            chart.add_data(s_ref, titles_from_data=True)
    else:
        raise ValueError(f"Unsupported chart type: {ctype}")

    if title:
        chart.title = title

    if ctype != "combo":
        d_minr, d_minc, d_maxr, d_maxc = _resolve_range(data_range)
        data_ref = Reference(ws, min_col=d_minc, min_row=d_minr,
                             max_col=d_maxc, max_row=d_maxr)
        chart.add_data(data_ref, titles_from_data=True)

    if cat_range:
        c_minr, c_minc, c_maxr, c_maxc = _resolve_range(cat_range)
        cat_ref = Reference(ws, min_col=c_minc, min_row=c_minr,
                            max_col=c_maxc, max_row=c_maxr)
        chart.set_categories(cat_ref)
    
    # Secondary axis support
    if secondary_axis and ctype != "combo":
        chart.y_axis.majorGridlines = None
    
    # Trendlines
    if spec.get("trendline"):
        for series in chart.series:
            series.trendline = Trendline(trendlineType=spec.get("trendline_type", "linear"))

    ws.add_chart(chart, anchor)


def _add_conditional_formatting(ws, spec: Dict[str, Any]) -> None:
    """Apply conditional formatting to a range."""
    cf_type = spec.get("type")
    rng = spec.get("range")
    if not rng:
        raise ValueError("conditional_formatting spec missing 'range'")
    
    r_minr, r_minc, r_maxr, r_maxc = _resolve_range(rng)
    
    if cf_type == "color_scale":
        start_color = spec.get("start_color", "FF0000")
        mid_color = spec.get("mid_color")
        end_color = spec.get("end_color", "00FF00")
        
        if mid_color:
            rule = ColorScaleRule(
                start_type="min", start_color=start_color,
                mid_type="percentile", mid_value=50, mid_color=mid_color,
                end_type="max", end_color=end_color
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color=start_color,
                end_type="max", end_color=end_color
            )
        ws.conditional_formatting.add(f"{rng}", rule)
    
    elif cf_type == "data_bar":
        color = spec.get("color", "0063B1")
        rule = DataBarRule(
            start_type="min", end_type="max",
            color=color
        )
        ws.conditional_formatting.add(f"{rng}", rule)
    
    elif cf_type == "icon_set":
        icon_style = spec.get("icon_style", "3TrafficLights1")
        rule = IconSetRule(
            icon_style=icon_style,
            type=spec.get("value_type", "percent"),
            values=spec.get("values", [0, 33, 67]),
            showValue=spec.get("show_value")
        )
        ws.conditional_formatting.add(f"{rng}", rule)
    
    elif cf_type == "cell_is":
        operator = spec.get("operator", "equal")
        formula = spec.get("formula")
        fill = spec.get("fill")
        font_color = spec.get("font_color")
        
        rule = CellIsRule(operator=operator, formula=[formula], stopIfTrue=True)
        if fill:
            rule.fill = PatternFill("solid", start_color=fill, end_color=fill)
        if font_color:
            rule.font = Font(color=font_color)
        ws.conditional_formatting.add(f"{rng}", rule)


def _add_data_validation(ws, spec: Dict[str, Any]) -> None:
    """Apply data validation to a range."""
    dv_type = spec.get("type")
    rng = spec.get("range")
    if not rng:
        raise ValueError("data_validation spec missing 'range'")
    dv = DataValidation(type=dv_type)
    if dv_type == "list":
        formula1 = spec.get("formula1", "")
        if formula1 and not formula1.startswith('"') and not formula1.startswith("="):
            formula1 = f'"{formula1}"'
        dv.formula1 = formula1
        dv.showDropDown = not bool(spec.get("show_dropdown", True))
    elif dv_type == "whole":
        dv.operator = spec.get("operator", "between")
        dv.formula1 = spec.get("formula1", "0")
        dv.formula2 = spec.get("formula2", "100")
    elif dv_type == "decimal":
        dv.operator = spec.get("operator", "between")
        dv.formula1 = spec.get("formula1", "0")
        dv.formula2 = spec.get("formula2", "1")
    elif dv_type == "textLength":
        dv.operator = spec.get("operator", "between")
        dv.formula1 = spec.get("formula1", "0")
        dv.formula2 = spec.get("formula2", "100")
    elif dv_type == "custom":
        dv.formula1 = spec.get("formula1", "")
    
    dv.error = spec.get("error", "Invalid value")
    dv.errorTitle = spec.get("error_title", "Error")
    dv.prompt = spec.get("prompt")
    dv.promptTitle = spec.get("prompt_title", "Input")
    
    ws.add_data_validation(dv)
    dv.add(rng)


def _build_sheet(wb: Workbook, sheet_spec: Dict[str, Any], default_font: Font,
                 first: bool) -> None:
    name = sheet_spec.get("name") or f"Sheet{len(wb.sheetnames) + 1}"
    if first and wb.sheetnames == ["Sheet"]:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    column_formats: Dict[int, str] = {}
    for col_spec in sheet_spec.get("columns", []) or []:
        col = col_spec.get("col")
        if not col:
            continue
        idx = column_index_from_string(col)
        if "width" in col_spec:
            ws.column_dimensions[col].width = float(col_spec["width"])
        if "number_format" in col_spec:
            column_formats[idx] = col_spec["number_format"]

    for r_idx, row_spec in enumerate(sheet_spec.get("rows", []) or [], start=1):
        if not isinstance(row_spec, dict):
            row_spec = {"values": list(row_spec)}
        _write_row(ws, r_idx, row_spec, default_font, column_formats)

    for merge in sheet_spec.get("merges", []) or []:
        ws.merge_cells(merge)

    freeze = sheet_spec.get("freeze")
    if freeze:
        ws.freeze_panes = freeze

    if sheet_spec.get("auto_width"):
        _auto_width(ws)

    for chart_spec in sheet_spec.get("charts", []) or []:
        _add_chart(ws, chart_spec)
    
    for cf_spec in sheet_spec.get("conditional_formatting", []) or []:
        _add_conditional_formatting(ws, cf_spec)
    
    for dv_spec in sheet_spec.get("data_validation", []) or []:
        _add_data_validation(ws, dv_spec)
    
    # Named ranges
    for nr_spec in sheet_spec.get("named_ranges", []) or []:
        nr_name = nr_spec.get("name")
        nr_range = nr_spec.get("range")
        if nr_name and nr_range:
            try:
                wb.create_named_range(nr_name, ws, nr_range)
            except Exception:
                # Named range might already exist, try to delete and recreate
                try:
                    if nr_name in wb.defined_names:
                        del wb.defined_names[nr_name]
                    wb.create_named_range(nr_name, ws, nr_range)
                except Exception:
                    pass  # Skip if still fails
    
    # Sheet protection
    protection = sheet_spec.get("protection")
    if protection:
        ws.protection = SheetProtection(
            sheet=protection.get("sheet", True),
            password=protection.get("password"),
            autoFilter=protection.get("autoFilter", True),
            deleteColumns=protection.get("deleteColumns", False),
            deleteRows=protection.get("deleteRows", False),
            formatCells=protection.get("formatCells", False),
            formatColumns=protection.get("formatColumns", False),
            formatRows=protection.get("formatRows", False),
            insertColumns=protection.get("insertColumns", False),
            insertRows=protection.get("insertRows", False),
            objects=protection.get("objects", False),
            pivotTables=protection.get("pivotTables", False),
            scenarios=protection.get("scenarios", False),
            selectLockedCells=protection.get("selectLockedCells", True),
            selectUnlockedCells=protection.get("selectUnlockedCells", True)
        )
    
    # Print settings
    print_settings = sheet_spec.get("print_settings")
    if print_settings:
        ws.page_setup.paperSize = print_settings.get("paperSize", 9)  # 9 = A4
        ws.page_setup.orientation = print_settings.get("orientation", "portrait")
        ws.page_setup.fitToPage = print_settings.get("fitToPage", True)
        ws.page_setup.fitToWidth = print_settings.get("fitToWidth", 1)
        ws.page_setup.fitToHeight = print_settings.get("fitToHeight", 0)
        
        if print_settings.get("print_area"):
            ws.print_area = print_settings["print_area"]
        
        if print_settings.get("repeat_rows"):
            ws.print_title_rows = print_settings["repeat_rows"]
        
        if print_settings.get("repeat_columns"):
            ws.print_title_cols = print_settings["repeat_columns"]


def build(spec: Dict[str, Any], output: Path, font_name: str, font_size: float) -> None:
    wb = Workbook()
    default_font = Font(name=font_name, size=font_size)

    sheets = spec.get("sheets") or []
    if not sheets:
        raise ValueError("spec.sheets must contain at least one sheet definition.")

    for i, sheet_spec in enumerate(sheets):
        _build_sheet(wb, sheet_spec, default_font, first=(i == 0))

    props = spec.get("properties") or {}
    if props.get("title"):
        wb.properties.title = props["title"]
    if props.get("creator"):
        wb.properties.creator = props["creator"]
    if props.get("subject"):
        wb.properties.subject = props["subject"]

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    p = argparse.ArgumentParser(description="Build an .xlsx workbook from a JSON spec.")
    p.add_argument("--spec", required=True, help="Path to spec JSON file")
    p.add_argument("--output", required=True, help="Output .xlsx path")
    p.add_argument("--font-name", default="Arial")
    p.add_argument("--font-size", type=float, default=11.0)
    args = p.parse_args()

    spec_path = Path(args.spec)
    if not spec_path.exists():
        print(f"Error: spec file not found: {spec_path}", file=sys.stderr)
        sys.exit(1)
    try:
        spec = json.loads(spec_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        print(f"Error: invalid JSON in spec: {exc}", file=sys.stderr)
        sys.exit(1)

    out = Path(args.output)
    if out.suffix.lower() != ".xlsx":
        print(f"Error: output must end with .xlsx (got {out.suffix})", file=sys.stderr)
        sys.exit(1)

    try:
        build(spec, out, font_name=args.font_name, font_size=args.font_size)
    except Exception as exc:
        print(f"Error building workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"Workbook built -> {out}  (run xlsx_recalc.py next to populate cached values)")


if __name__ == "__main__":
    main()
