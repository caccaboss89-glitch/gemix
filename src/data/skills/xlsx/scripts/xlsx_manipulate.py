#!/usr/bin/env python3
"""Combine / split / inspect xlsx workbooks.

Subcommands:
  merge          : take the first sheet of each input and stack them into one workbook
  extract-sheet  : copy a single sheet into a new workbook
  split          : write each sheet of an input as its own workbook
  info           : workbook metadata + sheet inventory (lighter than xlsx_inspect.py)

All operations preserve formulas (not cached values). Run xlsx_recalc.py on the
output if the user expects to open the file and see numbers immediately.
"""
import argparse
import json
import sys
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def _validate_xlsx(p: Path) -> Path:
    if not p.exists():
        raise FileNotFoundError(f"File not found: {p}")
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported extension: {p.suffix} (use .xlsx or .xlsm)")
    return p


def _safe_sheet_name(name: str, taken: set) -> str:
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\ , and unique."""
    cleaned = "".join(c for c in name if c not in "[]:*?/\\")[:31] or "Sheet"
    base = cleaned
    i = 2
    while cleaned in taken:
        suffix = f"_{i}"
        cleaned = (base[: 31 - len(suffix)] + suffix)
        i += 1
    taken.add(cleaned)
    return cleaned


def _copy_sheet(src_ws, dst_wb: Workbook, new_name: str):
    """Copy a sheet from one workbook to another, preserving values+formulas."""
    dst_ws = dst_wb.create_sheet(title=new_name)
    for row in src_ws.iter_rows(values_only=False):
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for r_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[r_idx].height = dim.height
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    return dst_ws


def cmd_merge(inputs: List[Path], output: Path, sheet_names: Optional[List[str]]) -> None:
    if not inputs:
        raise ValueError("No inputs provided.")
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    taken: set = set()
    for i, src in enumerate(inputs):
        wb = load_workbook(_validate_xlsx(src), data_only=False)
        first_sheet = wb[wb.sheetnames[0]]
        if sheet_names and i < len(sheet_names) and sheet_names[i]:
            preferred = sheet_names[i]
        else:
            preferred = src.stem
        target_name = _safe_sheet_name(preferred, taken)
        _copy_sheet(first_sheet, out_wb, target_name)

    output.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output)
    print(f"Merged {len(inputs)} workbook(s) -> {output}")


def cmd_extract_sheet(input_path: Path, sheet_name: str, output: Path) -> None:
    wb = load_workbook(_validate_xlsx(input_path), data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    _copy_sheet(wb[sheet_name], out_wb, _safe_sheet_name(sheet_name, set()))

    output.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output)
    print(f"Extracted sheet '{sheet_name}' -> {output}")


def cmd_split(input_path: Path, output_prefix: Path) -> Dict[str, Any]:
    wb = load_workbook(_validate_xlsx(input_path), data_only=False)
    output_prefix.parent.mkdir(parents=True, exist_ok=True)
    created: List[str] = []
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        _copy_sheet(wb[sheet_name], out_wb, _safe_sheet_name(sheet_name, set()))
        safe_part = "".join(c for c in sheet_name if c.isalnum() or c in "-_")[:40] or "sheet"
        out_path = output_prefix.parent / f"{output_prefix.name}_{i:03d}_{safe_part}.xlsx"
        out_wb.save(out_path)
        created.append(str(out_path))
        print(f"Sheet '{sheet_name}' -> {out_path}")
    return {
        "action": "split",
        "input": str(input_path),
        "outputs": created,
        "count": len(created),
    }


def cmd_info(input_path: Path) -> None:
    wb = load_workbook(_validate_xlsx(input_path), data_only=True)
    props = wb.properties
    info = {
        "file": str(input_path),
        "properties": {
            "title": props.title,
            "creator": props.creator,
            "subject": props.subject,
            "modified": str(props.modified) if props.modified else None,
        },
        "sheets": [
            {
                "name": name,
                "max_row": wb[name].max_row,
                "max_col": wb[name].max_column,
                "max_col_letter": get_column_letter(wb[name].max_column or 1),
            }
            for name in wb.sheetnames
        ],
    }
    print(json.dumps(info, indent=2, ensure_ascii=False, default=str))


def main() -> None:
    parser = argparse.ArgumentParser(description="xlsx merge / extract-sheet / split / info")
    sub = parser.add_subparsers(dest="action", required=True)

    m = sub.add_parser("merge", help="Merge first sheets of inputs into one workbook")
    m.add_argument("--inputs", nargs="+", required=True)
    m.add_argument("--output", required=True)
    m.add_argument("--sheet-names",
                   help="Comma-separated names for the merged sheets (matches --inputs order)")

    e = sub.add_parser("extract-sheet", help="Copy a single sheet to a new workbook")
    e.add_argument("--input", required=True)
    e.add_argument("--sheet", required=True)
    e.add_argument("--output", required=True)

    s = sub.add_parser("split", help="Write each sheet to its own workbook")
    s.add_argument("--input", required=True)
    s.add_argument("--output-prefix", required=True)

    i = sub.add_parser("info", help="Print metadata + sheet inventory")
    i.add_argument("--input", required=True)

    args = parser.parse_args()
    try:
        if args.action == "merge":
            sheet_names = args.sheet_names.split(",") if args.sheet_names else None
            cmd_merge(
                [Path(p) for p in args.inputs],
                Path(args.output),
                sheet_names,
            )
        elif args.action == "extract-sheet":
            cmd_extract_sheet(Path(args.input), args.sheet, Path(args.output))
        elif args.action == "split":
            result = cmd_split(Path(args.input), Path(args.output_prefix))
            print(json.dumps(result, indent=2))
        elif args.action == "info":
            cmd_info(Path(args.input))
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
