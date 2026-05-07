#!/usr/bin/env python3
"""Inspect an .xlsx/.xlsm workbook and emit a JSON summary.

Designed to be the FIRST tool the AI runs on any pre-existing spreadsheet,
because the GemiX runtime does not auto-parse spreadsheets the way it does PDFs.
"""
import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


def _coerce(value: Any) -> Any:
    """Make a cell value JSON-safe."""
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    # datetime, date, time, Decimal, etc.
    return str(value)


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _scan_sheet(ws, rows_sample: int, scan_formulas_ws=None) -> Dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    headers: List[Any] = []
    if max_row >= 1:
        for col in range(1, max_col + 1):
            headers.append(_coerce(ws.cell(row=1, column=col).value))

    sample_rows: List[Dict[str, Any]] = []
    if rows_sample > 0 and max_row >= 2:
        last = min(max_row, 1 + rows_sample)
        for r in range(2, last + 1):
            row_vals = [_coerce(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
            sample_rows.append({"row": r, "values": row_vals})

    # Errors are scanned in the data_only view (cached values).
    errors: Dict[str, List[str]] = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v in ERROR_TOKENS:
                errors.setdefault(v, []).append(f"{ws.title}!{cell.coordinate}")

    # Formulas must be counted from a non-data_only view.
    formula_count = 0
    if scan_formulas_ws is not None:
        for row in scan_formulas_ws.iter_rows():
            for cell in row:
                if _is_formula(cell.value):
                    formula_count += 1

    merged = [str(rng) for rng in ws.merged_cells.ranges]

    return {
        "name": ws.title,
        "max_row": max_row,
        "max_col": max_col,
        "max_col_letter": get_column_letter(max_col) if max_col else "",
        "headers": headers,
        "formula_count": formula_count,
        "merged_cells": merged,
        "errors": errors,
        "sample_rows": sample_rows,
    }


def inspect(
    input_path: Path,
    rows_sample: int,
    sheets_filter: Optional[List[str]],
    data_only: bool,
) -> Dict[str, Any]:
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported extension: {input_path.suffix} (use .xlsx or .xlsm)")

    wb_values = load_workbook(input_path, data_only=data_only, read_only=False)
    # Second view to count formulas regardless of data_only setting.
    wb_formulas = wb_values
    if data_only:
        wb_formulas = load_workbook(input_path, data_only=False, read_only=False)

    sheets_out: List[Dict[str, Any]] = []
    target_names = [s.strip() for s in sheets_filter] if sheets_filter else wb_values.sheetnames
    for name in target_names:
        if name not in wb_values.sheetnames:
            sheets_out.append({"name": name, "error": "sheet not found"})
            continue
        ws_v = wb_values[name]
        ws_f = wb_formulas[name] if name in wb_formulas.sheetnames else None
        sheets_out.append(_scan_sheet(ws_v, rows_sample, scan_formulas_ws=ws_f))

    props = wb_values.properties
    return {
        "file": str(input_path),
        "data_only": data_only,
        "sheet_names": wb_values.sheetnames,
        "properties": {
            "title": props.title,
            "creator": props.creator,
            "subject": props.subject,
            "modified": str(props.modified) if props.modified else None,
        },
        "sheets": sheets_out,
    }


def main() -> None:
    p = argparse.ArgumentParser(description="Inspect an .xlsx workbook (JSON summary).")
    p.add_argument("--input", required=True)
    p.add_argument("--rows-sample", type=int, default=5)
    p.add_argument("--sheets", help="Comma-separated subset of sheet names")
    p.add_argument("--no-data-only", action="store_true",
                   help="Read formula strings instead of cached values")
    p.add_argument("--output", help="Write JSON here (default: stdout)")
    args = p.parse_args()

    sheets_filter = args.sheets.split(",") if args.sheets else None
    try:
        report = inspect(
            Path(args.input),
            rows_sample=max(0, args.rows_sample),
            sheets_filter=sheets_filter,
            data_only=not args.no_data_only,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    payload = json.dumps(report, indent=2, ensure_ascii=False, default=str)
    if args.output:
        out = Path(args.output)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(payload, encoding="utf-8")
        print(f"Inspection written -> {out}")
    else:
        print(payload)


if __name__ == "__main__":
    main()
