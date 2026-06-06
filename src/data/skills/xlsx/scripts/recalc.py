"""
Excel formula recalculation script.

openpyxl writes formulas as strings but never computes their results, so a file
it produced has empty/last-cached values until something opens it. This script
drives headless LibreOffice to recalculate every formula in every sheet, save
the file back, then scans all cells for Excel errors and reports them as JSON.

Usage:
    python /skills/xlsx/scripts/recalc.py <excel_file> [timeout_seconds]
"""

import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from soffice import get_soffice_env

MACRO_DIR = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def setup_libreoffice_macro():
    macro_dir = os.path.expanduser(MACRO_DIR)
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    if not os.path.exists(macro_dir):
        # First-run init: let LibreOffice create its profile tree.
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=60,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "timeout",
        str(timeout),
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    # returncode 124 = the `timeout` wrapper fired; LibreOffice may still have
    # saved before being killed, so we continue to the error scan either way.
    if result.returncode not in (0, 124):
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)
        error_details = {err: [] for err in EXCEL_ERRORS}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in EXCEL_ERRORS:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()
        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice.")
        print("Returns JSON: status, total_errors, total_formulas, error_summary.")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    print(json.dumps(recalc(filename, timeout), indent=2))


if __name__ == "__main__":
    main()
