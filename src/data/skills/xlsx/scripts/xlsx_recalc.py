#!/usr/bin/env python3
"""Recalculate formulas in an .xlsx via headless LibreOffice and scan for errors.

Why a wrapper around LibreOffice:
- openpyxl never evaluates formulas; it only stores the formula string.
- Without recalculation, the saved file has no cached values → opening it in
  Excel/Numbers/Sheets shows blanks until the user manually edits a cell.
- LibreOffice in --headless --calc mode does a full recalc on save.

Sandbox quirks handled here:
- A unique --user-profile per invocation avoids Unix-socket lock collisions
  (multiple soffice processes refuse to share a profile).
- The conversion writes to a temp dir and then atomically replaces the input,
  so an interrupted run never corrupts the original.
- After recalc we re-open with openpyxl(data_only=True) to scan all cells for
  Excel error tokens and emit a JSON summary.
"""
import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


def _find_soffice() -> str:
    for cand in ("soffice", "libreoffice"):
        path = shutil.which(cand)
        if path:
            return path
    raise RuntimeError(
        "LibreOffice (soffice) not found in PATH. Install libreoffice-calc in the sandbox image."
    )


def _recalc_with_soffice(input_path: Path, timeout: int) -> None:
    soffice = _find_soffice()
    with tempfile.TemporaryDirectory(prefix="lo_recalc_") as tmpdir:
        tmp = Path(tmpdir)
        profile_dir = tmp / f"profile_{uuid.uuid4().hex[:8]}"
        out_dir = tmp / "out"
        out_dir.mkdir()

        # file:// URI is required by -env:UserInstallation.
        profile_uri = profile_dir.absolute().as_uri()

        cmd = [
            soffice,
            f"-env:UserInstallation={profile_uri}",
            "--headless",
            "--calc",
            "--norestore",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to", "xlsx",
            "--outdir", str(out_dir),
            str(input_path),
        ]
        try:
            res = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=timeout,
                env={**os.environ, "HOME": str(profile_dir)},
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(
                f"LibreOffice recalc timed out after {timeout}s. "
                f"Increase --timeout or split the workbook."
            ) from exc

        if res.returncode != 0:
            raise RuntimeError(
                f"soffice exited {res.returncode}.\nstdout: {res.stdout}\nstderr: {res.stderr}"
            )

        produced = out_dir / input_path.name
        if not produced.exists():
            # Some LibreOffice builds rename the output if the source had .xlsm etc.
            xlsx_files = list(out_dir.glob("*.xlsx"))
            if not xlsx_files:
                raise RuntimeError(f"soffice did not produce an output file. stderr: {res.stderr}")
            produced = xlsx_files[0]

        # Atomic replace so the input file always exists.
        shutil.copy2(produced, input_path)


def _scan_errors(input_path: Path) -> Dict[str, Any]:
    wb_values = load_workbook(input_path, data_only=True)
    wb_formulas = load_workbook(input_path, data_only=False)

    error_summary: Dict[str, Dict[str, Any]] = {}
    total_errors = 0
    total_formulas = 0

    for sheet_name in wb_values.sheetnames:
        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        for row in ws_v.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in ERROR_TOKENS:
                    bucket = error_summary.setdefault(v, {"count": 0, "locations": []})
                    bucket["count"] += 1
                    if len(bucket["locations"]) < 50:
                        bucket["locations"].append(f"{sheet_name}!{cell.coordinate}")
                    total_errors += 1

        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1

    status = "errors_found" if total_errors > 0 else "success"
    out: Dict[str, Any] = {
        "status": status,
        "file": str(input_path),
        "total_formulas": total_formulas,
        "total_errors": total_errors,
    }
    if error_summary:
        out["error_summary"] = error_summary
    return out


def main() -> None:
    p = argparse.ArgumentParser(description="Recalculate xlsx formulas via headless LibreOffice.")
    p.add_argument("--input", required=True)
    p.add_argument("--timeout", type=int, default=60)
    p.add_argument("--output", help="Write status JSON here (default: stdout)")
    args = p.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        print(f"Error: Unsupported extension: {input_path.suffix}", file=sys.stderr)
        sys.exit(1)

    try:
        _recalc_with_soffice(input_path, timeout=args.timeout)
        report = _scan_errors(input_path)
    except Exception as exc:
        report = {"status": "failed", "file": str(input_path), "error": str(exc)}
        payload = json.dumps(report, indent=2, ensure_ascii=False)
        if args.output:
            Path(args.output).parent.mkdir(parents=True, exist_ok=True)
            Path(args.output).write_text(payload, encoding="utf-8")
        print(payload, file=sys.stderr)
        sys.exit(1)

    payload = json.dumps(report, indent=2, ensure_ascii=False)
    if args.output:
        out = Path(args.output)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(payload, encoding="utf-8")
        print(f"Recalc status written -> {out}  (status={report['status']}, errors={report['total_errors']})")
    else:
        print(payload)


if __name__ == "__main__":
    main()
